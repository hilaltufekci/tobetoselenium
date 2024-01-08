from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
import pytest
import openpyxl
from constants import globalConstants as c

class Test_CaseHomework:

    def setup_method(self):
        self.driver=webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window()

    def teardown_method(self): 
        self.driver.quit()
    
    def getData():
        excel = openpyxl.load_workbook("data/invalid_login.xlsx")
        sheet = excel["Sayfa1"] 
        rows = sheet.max_row 
        data = []
        for i in range(2,rows+1):
            emailname = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((emailname,password))

        return data

    @pytest.mark.parametrize("emailname,password",getData())
    def test_invalid(self,emailname,password):
        emailnameInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.EMAİLNAME_XPATH))).send_keys(emailname)
        passwordInput=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,c.PASSWORD_XPATH))).send_keys(password)
        loginButton=self.driver.find_element(By.XPATH,c.LOGIN_BUTTON_XPATH).click()
        errormessage=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='__next']/div/main/div[2]/div/div[2]")))
        assert errormessage.text=="• Geçersiz e-posta veya şifre."
        sleep(5)